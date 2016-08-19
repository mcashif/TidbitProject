
from .forms import DocumentForm
from .models import ExcelFile,XMLData
from django.shortcuts import render_to_response
from django.template import RequestContext
from django.conf import settings
import openpyxl
import unicodedata
import h5py as hdf
from django.utils.encoding import smart_str
from django.template import loader
from django.http import Http404, HttpResponse
from django.views.decorators.csrf import csrf_exempt
import json
import os
import stat
import shutil
from lxml import etree
import shutil
import re
import io
from .serializers import XMLSerializer
from rest_framework import routers, serializers, viewsets
from rest_framework.parsers import MultiPartParser,FileUploadParser,FormParser
from rest_framework.response import Response





#Global Data
listofdData=[]
sheetList=[]
textNodes=[]
dbNodes=[]
xmlPath=""

#Helper Function Remove Directory
def _remove_readonly(fn, path_, excinfo):
    # Handle read-only files and directories
    if fn is os.rmdir:
        os.chmod(path_, stat.S_IWRITE)
        os.rmdir(path_)
    elif fn is os.remove:
        os.lchmod(path_, stat.S_IWRITE)
        os.remove(path_)


def force_remove_file_or_symlink(path_):
    try:
        os.remove(path_)
    except OSError:
        os.lchmod(path_, stat.S_IWRITE)
        os.remove(path_)


# Code from shutil.rmtree()
def is_regular_dir(path_):
    try:
        mode = os.lstat(path_).st_mode
    except os.error:
        mode = 0
    return stat.S_ISDIR(mode)


def clear_dir(path_):
    if is_regular_dir(path_):
        # Given path is a directory, clear its content
        for name in os.listdir(path_):
            fullpath = os.path.join(path_, name)
            if is_regular_dir(fullpath):
                shutil.rmtree(fullpath, onerror=_remove_readonly)
            else:
                force_remove_file_or_symlink(fullpath)
    else:
        # Given path is a file or a symlink.
        # Raise an exception here to avoid accidentally clearing the content
        # of a symbolic linked directory.
        raise OSError("Cannot call clear_dir() on a symbolic link")

#Helper Functions Excel Data
def isCellEmpty(cell):

    if cell.value is None:
        return True
    else:
        if(isCellEmptSpaces(cell)):
            return True

    return False

def isCellEmptSpaces(cell):
    val=str(cell.value)
    if(val==' '):
        return True

    return False

def isCellFormula(cell):

    val = str(cell.value)
    if(val.startswith('=')):
        return True

    return False


def is_number(s):
    try:
        float(s)
        return True
    except ValueError:
        pass

    try:
        unicodedata.numeric(s)
        return True
    except (TypeError, ValueError):
        pass

    return False

def isCellNumber(cell):
    val = str(cell.value)
    if(is_number(val)==True):
        return True

    return False

def isCellConnected(sheet,col,row):

    hr=sheet.get_highest_row()+1
    hc=sheet.get_highest_column()+1

    #check Right if value return True
    if(col!=hc):
        if(isCellEmpty(sheet.cell(column=col+1, row=row))==False):
            return True

    #check Below if value return True
    if(row!=hr):
        if(isCellEmpty(sheet.cell(column=col, row=row+1))==False):
            return True

    #check Left,Boundry Column Elements, if value return True
    if(col>1):
        if(isCellEmpty(sheet.cell(column=col-1, row=row))==False):
            return True

    #check top,Boundry Column Elements, if value return True
    if(row>1):
        if(isCellEmpty(sheet.cell(column=col, row=row-1))==False):
            return True


    return False


def findTopVaiable(sheet, cell, col, row):

    for row in range(row,0,-1):
        if(isCellEmpty(sheet.cell(column=col, row=row))==True):
            continue
        value=str(sheet.cell(column=col, row=row).value)
        cell=sheet.cell(column=col, row=row)
        if(isCellFormula(cell)==False):
            if(is_number(value)==False):
                return value

    return "No Top Heading"


def findLeftVaiable(sheet, cell, col, row):

    for col in range(col,0,-1):
        if(isCellEmpty(sheet.cell(column=col, row=row))==True):
            continue
        value=str(sheet.cell(column=col, row=row).value)
        cell=sheet.cell(column=col, row=row)
        if(isCellFormula(cell)==False):
            if(is_number(value)==False):
                return value

    return "No Left Heading"

def cellNumberUsedInSheet(sheet,cell,colv,rowv):

    for row in sheet.rows:
        for cellRow in row:
            cellValue=str(cellRow.value)
            cellCordinate=str(cell.coordinate)
            if ( cellValue.find(cellCordinate)>-1 & cellValue.find("!")==-1):
                return "("+str(cell.value)+")"+"(Intermediate Cell)" + "(Used at:"+cellRow.coordinate+")"+"("+findTopVaiable(sheet,cell,colv,rowv)+")"+"("+findLeftVaiable(sheet,cell,colv,rowv)+")"

    return "("+str(cell.value)+")"+"(Output)" +"("+findTopVaiable(sheet,cell,colv,rowv)+")"+"("+findLeftVaiable(sheet,cell,colv,rowv)+")"

def cellUsedInSheet(sheet,cell,formulaOutput,colv,rowv):

    for row in sheet.rows:
        for cellRow in row:
            cellValue=str(cellRow.value)
            cellCordinate=str(cell.coordinate)
            if ( cellValue.find(cellCordinate)>-1 & cellValue.find("!")==-1):
                return "("+str(cell.value)+")"+"("+formulaOutput+")"+"(Intermediate Equation)" + "(Used at:"+cellRow.coordinate+")"+"("+findTopVaiable(sheet,cell,colv,rowv)+")"+"("+findLeftVaiable(sheet,cell,colv,rowv)+")"

    return "("+str(cell.value)+")"+"("+formulaOutput+")"+"(Output)" +"("+findTopVaiable(sheet,cell,colv,rowv)+")"+"("+findLeftVaiable(sheet,cell,colv,rowv)+")"


def getFormulaType(sheet,cell,col,row,formulaOutput):

    valStr=str(cell.value)
    valCordinate=str(cell.coordinate)

    if "!" in valStr:
        return "("+valStr+")"+"("+formulaOutput+")"+"(input)"+"("+findTopVaiable(sheet,cell,col,row)+")"+"("+findLeftVaiable(sheet,cell,col,row)+")"


    return cellUsedInSheet(sheet,cell,formulaOutput,col,row)

# End Helper Function

#Process Each Sheet and Get Data
def extractDataFromSheet(sheetformula,sheetvalue,col,row,sheet):

#Return if Empty
    if(isCellEmpty(sheetformula.cell(column=col, row=row))):
        return 0


    value=str(sheetvalue.cell(column=col, row=row).value)
    formula=str(sheetformula.cell(column=col, row=row).value)

    if(isCellNumber(sheetformula.cell(column=col, row=row))):
        val=[0,sheet,sheetformula.cell(column=col, row=row).coordinate,value,findTopVaiable(sheetvalue,sheetvalue.cell(column=col, row=row),col,row),findLeftVaiable(sheetvalue,sheetvalue.cell(column=col, row=row),col,row)]
        return val
    else:
        if(isCellFormula(sheetformula.cell(column=col, row=row))):
            val=[1,sheet,sheetformula.cell(column=col, row=row).coordinate,formula,value,findTopVaiable(sheetvalue,sheetvalue.cell(column=col, row=row),col,row),findLeftVaiable(sheetvalue,sheetvalue.cell(column=col, row=row),col,row)]

            return val


    return 0

#Process Workbook
def processWorkBook(path):

    #workbook loaded but it display formulas instead of values of formulas so we need two loaded object one with formula and one with value, both work side by side
    workBook = openpyxl.load_workbook(settings.PROJECT_ROOT+"/media/"+path)
    #value object
    workBookValued = openpyxl.load_workbook(settings.PROJECT_ROOT+"/media/"+path,data_only=True)

    workBookSheets=workBook.get_sheet_names()

    listofdData[:] = []
    sheetList[:] = []

    #for each sheet we process
    for sheet in workBookSheets:

        workSheetFormula = workBook.get_sheet_by_name(sheet)
        workSheetValued = workBookValued.get_sheet_by_name(sheet)
        hr=workSheetFormula.get_highest_row()
        hc=workSheetFormula.get_highest_column()
        sheetList.append(sheet)

        for row in range(1, hr+1):
            for col in range(1, hc+1):
                val=extractDataFromSheet(workSheetFormula,workSheetValued,col,row,sheet)
                if(val!=0):
                    listofdData.append(val)


    return settings.PROJECT_ROOT+"/media/documents/data.hdf5"


#ajax call to send Edittied data and make hdf5 file
@csrf_exempt
def makeHdfFile(request):


    if request.is_ajax() and request.POST:

        data = request.POST.get("data")
        jObject = json.loads(data)

        outfile = hdf.File(settings.PROJECT_ROOT+"/media/documents/data.hdf5",'w')

        fr="formula"
        va="value"
        tp="top-label"
        lf="left-label"
        sheet="non"

        g1=0
        g2=0
        g3=0
        g4=0

        for jo in jObject:

            if(sheet==jo['sheet']):
                fr1=jo['cid']+":"+jo['formula']
                va1=jo['cid']+":"+jo['value']
                tp1=jo['cid']+":"+jo['topLabel']
                lf1=jo['cid']+":"+jo['leftLabel']
                dset=g1.create_dataset(fr1, data=jo['cid']+":"+jo['formula'])
                dset=g2.create_dataset(va1, data=jo['cid']+":"+jo['value'])
                dset=g3.create_dataset(tp1, data=jo['cid']+":"+jo['topLabel'])
                dset=g4.create_dataset(lf1, data=jo['cid']+":"+jo['leftLabel'])

            if(sheet=="non"):
                sheet=jo['sheet']
                grp_sheet = outfile.create_group(sheet)
                g1=grp_sheet.create_group(fr)
                g2=grp_sheet.create_group(va)
                g3=grp_sheet.create_group(tp)
                g4=grp_sheet.create_group(lf)
                fr1=jo['cid']+":"+jo['formula']
                va1=jo['cid']+":"+jo['value']
                tp1=jo['cid']+":"+jo['topLabel']
                lf1=jo['cid']+":"+jo['leftLabel']
                dset=g1.create_dataset(fr1, data=jo['cid']+":"+jo['formula'])
                dset=g2.create_dataset(va1, data=jo['cid']+":"+jo['value'])
                dset=g3.create_dataset(tp1, data=jo['cid']+":"+jo['topLabel'])
                dset=g4.create_dataset(lf1, data=jo['cid']+":"+jo['leftLabel'])

            if(sheet!=jo['sheet']):
                sheet=jo['sheet']
                grp_sheet = outfile.create_group(sheet)
                g1=grp_sheet.create_group(fr)
                g2=grp_sheet.create_group(va)
                g3=grp_sheet.create_group(tp)
                g4=grp_sheet.create_group(lf)
                fr1=jo['cid']+":"+jo['formula']
                va1=jo['cid']+":"+jo['value']
                tp1=jo['cid']+":"+jo['topLabel']
                lf1=jo['cid']+":"+jo['leftLabel']
                dset=g1.create_dataset(fr1, data=jo['cid']+":"+jo['formula'])
                dset=g2.create_dataset(va1, data=jo['cid']+":"+jo['value'])
                dset=g3.create_dataset(tp1, data=jo['cid']+":"+jo['topLabel'])
                dset=g4.create_dataset(lf1, data=jo['cid']+":"+jo['leftLabel'])




        outfile.close()


        filename = settings.PROJECT_ROOT+"/media/documents/data.hdf5"
        response = HttpResponse(content_type='application/force-download')
        response['Content-Disposition'] = 'attachment; filename=%s' % smart_str("data.hdf5")
        response['X-Sendfile'] = smart_str(filename)
        # It's usually a good idea to set the 'Content-Length' header too.
        # You can also set any other required headers: Cache-Control, etc.
        return response

    else:
        raise Http404



#Start Point
def index(request):

    # Handle file upload
    if request.method == 'POST':
        form = DocumentForm(request.POST, request.FILES)
        if form.is_valid():

            #CLear All Old Data in Database and directory
            ExcelFile.objects.all().delete()
            clear_dir(settings.PROJECT_ROOT+"/media/documents/")
            #////////////////////////////////////////////////////
            #Read Excel and load into Database for processing, direct uploading can be done to directory but with database we can have record if needed
            newdoc = ExcelFile(docfile = request.FILES['docfile'])
            newdoc.save()
            #entry point to processing of file
            processWorkBook(newdoc.docfile.name)
            template = loader.get_template('tidbit/index.html')
            context = {

                'listofdData': listofdData,
                'sheetList': sheetList,
            }

            return HttpResponse(template.render(context, request))
    else:
        form = DocumentForm() # A empty, unbound form


    # Render GUI
    return render_to_response(
        'tidbit/index.html',
        {'form': form},
        context_instance=RequestContext(request)
    )

#Start Point
def index2(request):

    # Handle file upload
    if request.method == 'POST':
        form = DocumentForm(request.POST, request.FILES)
        if form.is_valid():

            #CLear All Old Data in Database and directory
            ExcelFile.objects.all().delete()
            clear_dir(settings.PROJECT_ROOT+"/media/documents/")
            #////////////////////////////////////////////////////
            #Read Excel and load into Database for processing, direct uploading can be done to directory but with database we can have record if needed
            newdoc = ExcelFile(docfile = request.FILES['docfile'])
            newdoc.save()
            #entry point to processing of file

            documents=ExcelFile.objects.all();


            template = loader.get_template('tidbit/index2.html')
            context = {

                'documents': documents,
            }

            return HttpResponse(template.render(context, request))
    else:
        form = DocumentForm() # A empty, unbound form


    # Render GUI
    return render_to_response(
        'tidbit/index2.html',
        {'form': form},
        context_instance=RequestContext(request)
    )

    return HttpResponse(template.render(context, request))


#Start Point
def index3(request):

    # Handle file upload
    if request.method == 'POST':
        form = DocumentForm(request.POST, request.FILES)
        if form.is_valid():

            #CLear All Old Data in Database and directory
            ExcelFile.objects.all().delete()
            clear_dir(settings.PROJECT_ROOT+"/media/documents/")
            #////////////////////////////////////////////////////
            #Read Excel and load into Database for processing, direct uploading can be done to directory but with database we can have record if needed
            newdoc = ExcelFile(docfile = request.FILES['docfile'])
            newdoc.save()
            #entry point to processing of file

            documents=ExcelFile.objects.all();


            template = loader.get_template('tidbit/index3.html')
            context = {

                'documents': documents,
            }

            return HttpResponse(template.render(context, request))
    else:
        form = DocumentForm() # A empty, unbound form


    # Render GUI
    return render_to_response(
        'tidbit/index3.html',
        {'form': form},
        context_instance=RequestContext(request)
    )

    return HttpResponse(template.render(context, request))



#Start Point
def index4(request):
    template = loader.get_template("tidbit/index4.html")
    return HttpResponse(template.render())


#Start Point
def index5(request):
    template = loader.get_template("tidbit/index5.html")
    return HttpResponse(template.render())








def elements_equal(e1, e2):

    if type(e1) != type(e2):
        return False
    if e1.tag != e1.tag: return False
    if e1.text != e2.text: return False
    if e1.tail != e2.tail: return False
    if e1.attrib != e2.attrib: return False
    if len(e1) != len(e2): return False
    return all([elements_equal(c1, c2) for c1, c2 in zip(e1, e2)])





#////////////////XML////////////////////////////////////////////////////

def isendNode(node):

    if(node.getchildren()):
      return False
    else:
      return True

def isendNodeWithText(node):

    if(node.getchildren()):
      return False
    else:
        return (node.text and node.text.strip()) or None


def makeXMLFromNode(node):

    xmlList=[]
    nodeX=node

    while etree.iselement(nodeX):
             nodeX=nodeX.getparent()
             if etree.iselement(nodeX):
                 xmlList.append(nodeX)

    return xmlList

def makeXMLFromNodeStr(node):

    strList=node.tag
    nodeX=node

    while etree.iselement(nodeX):
             nodeX=nodeX.getparent()
             if etree.iselement(nodeX):
                 strList=strList+"->"+nodeX.tag

    return strList

def getParentID(node):

    for i in range(0, len(dbNodes)):
        if(dbNodes[i][0]==node.getparent()):
            return dbNodes[i][1]

    return 0

def getTxtData(el):

    txtData=""

    if(el.text):
        txtData=el.text
    else:
        if(el.tail):
            txtData=el.tail

    return txtData

def populateDB(path):

    parent="Root"
    idParent=0
    dbNodes[:] = []


    XMLData.objects.all().delete()
    parser = etree.XMLParser(load_dtd= True)
    tree = etree.parse(settings.PROJECT_ROOT+"/media/"+path,parser)
    for el in tree.iter():
             if etree.iselement(el):

                if etree.iselement(el.getparent()):
                    parent= el.getparent().tag
                    idParent=getParentID(el)

                link=tree.getpath(el)
                txtData=getTxtData(el)
                newRecord = XMLData(nodeName = el.tag, nodeparentName = parent, nodeparentCode = idParent ,nodeattribute = str(dict(el.attrib)), nodedata =  txtData, linktoparent= link)
                newRecord.save()
                dbNodes.append([el,newRecord.id])
                print("OK")





def readNodes(path):
    xmlPath=path

    utf8_parser = etree.XMLParser(load_dtd= True)

    tree = etree.parse(settings.PROJECT_ROOT+"/media/"+path,utf8_parser)


    textNodes[:] = []

    for el in tree.iter():
        if(isendNodeWithText(el)):
            d = dict(el.attrib)
            textNodes.append([el,el.getparent(),el.tag,el.text,d,el.tag+"-"+ str(d)])

    return textNodes


#Start Point
@csrf_exempt
def makexml(request):

    if request.is_ajax() and request.POST:

        post_text = request.POST.get('nodeIndex')
        response_data = {}

        result=makeXMLFromNode(textNodes[int(post_text)][0])

        strResult=""

        for i in range(0, len(result)):
            strResult=strResult+"["+result[i].tag+"]"


        response_data['result'] = strResult


        return HttpResponse(
            json.dumps(response_data),
            content_type="application/json"
        )
    else:
        return HttpResponse(
            json.dumps({"nothing to see": "this isn't happening"}),
            content_type="application/json"
        )

def cleanXML2(path):
            num=""
            with io.open(settings.PROJECT_ROOT+"/media/"+path, "r", encoding="utf-8") as my_file:
                  my_unicode_string = my_file.read()
                  num = re.sub(r'<QOis4/>', "||", my_unicode_string)


            with io.open(settings.PROJECT_ROOT+"/media/"+path, "w", encoding="utf-8") as my_file:
                 my_file.write(num)


def getRealNext(node):


    for el in node.findall('.//'):
        if el==node:
            continue
        if etree.iselement(el):
           return el

    return False

def cleanXMLLevel2(path):

    parser = etree.XMLParser(load_dtd= True)
    tree = etree.parse(settings.PROJECT_ROOT+"/media/"+path,parser)


    for el in tree.iter():
             if etree.iselement(el):
                    nextReal= getRealNext(el)
                    if etree.iselement(nextReal) and isendNode(nextReal):
                        strVal=str(etree.tostring(nextReal))
                        if "/>" in strVal:
                            tag=nextReal.tag
                            strText=el.text
                            el.text=""
                            sub_node_element = etree.XML("<" + tag + "/>")
                            el.insert(0,sub_node_element)
                            sub_node_element.tail=strText


    tree.write(settings.PROJECT_ROOT+"/media/"+path, pretty_print=True)


def cleanXMLLevel1(path):

    parser = etree.XMLParser(load_dtd= True)
    tree = etree.parse(settings.PROJECT_ROOT+"/media/"+path,parser)


    for el in tree.iter():
             if etree.iselement(el):
                if el.text==":" :
                    if isendNode(el):
                        el.getparent().text=el.getparent().text+":"
                        el.getparent().remove(el)


    tree.write(settings.PROJECT_ROOT+"/media/"+path, pretty_print=True)

#Start Point
def index7(request):
    # Handle file upload
    if request.method == 'POST':
        form = DocumentForm(request.POST, request.FILES)
        if form.is_valid():

            #CLear All Old Data in Database and directory
            ExcelFile.objects.all().delete()
            clear_dir(settings.PROJECT_ROOT+"/media/documents/")
            #////////////////////////////////////////////////////
            #Read Excel and load into Database for processing, direct uploading can be done to directory but with database we can have record if needed
            newdoc = ExcelFile(docfile = request.FILES['docfile'])
            newdoc.save()
            #entry point to processing of file

            cleanXMLLevel1(newdoc.docfile.name)
            cleanXMLLevel2(newdoc.docfile.name)

            populateDB(newdoc.docfile.name)

            documents=ExcelFile.objects.all();

            #readNodes(newdoc.docfile.name)

            template = loader.get_template('tidbit/index7.html')
            context = {
                'textNodes' : textNodes,
                'documents': documents,
            }

            return HttpResponse(template.render(context, request))
    else:
        form = DocumentForm() # A empty, unbound form


    # Render GUI
    return render_to_response(
        'tidbit/index7.html',
        {'form': form},
        context_instance=RequestContext(request)
    )

    return HttpResponse(template.render(context, request))

def treeview(request):
        # Render GUI
        return render_to_response(
            'tidbit/tree.html', context_instance=RequestContext(request)
        )


# ViewSets define the view behavior.
class XMLViewSet(viewsets.ModelViewSet):

        queryset = ExcelFile.objects.all()
        serializer_class = XMLSerializer
        parser_classes = (MultiPartParser,FormParser,)
        def put(self, request, format=None):

            #CLear All Old Data in Database and directory
            ExcelFile.objects.all().delete()
            #////////////////////////////////////////////////////
            #Read Excel and load into Database for processing, direct uploading can be done to directory but with database we can have record if needed
            newdoc = ExcelFile(docfile = request.FILES['docfile'])
            newdoc.save()

            cleanXMLLevel1(newdoc.docfile.name)
            cleanXMLLevel2(newdoc.docfile.name)

            populateDB(newdoc.docfile.name)

            # ...
            # do some stuff with uploaded file
            # ...


            return Response(status=202)
